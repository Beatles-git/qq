#执行文件

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="gb18030")
from QCD.qcd_requests import qcd_requests#从当前文件夹下的含有请求模块的工作簿引入请求函数
from QCD.qcd_test_case import test_case#从当前文件夹下的含有测试数据函数的工作簿引入调用测试数据的函数
def run_write(file_name,sheet_name,row,column_1,column_2):
    all_case=test_case(file_name,sheet_name)#调用引入测试数据的函数，将返回的结果存放在all_case中
    token_1 = ''
    for i in range(len(all_case)):
        result_01 = {}
        ip='http://120.78.128.25:8766'
        uri=all_case[i][3]
        url=ip+uri
        method=all_case[i][4]
        param=eval(all_case[i][5])
        print(all_case[i])
        if sheet_name in ['register','login']:
            result = qcd_requests(url, param, method, token=None)  # 发起登录请求，将返回的值存放在result中
            print(result)
        elif sheet_name in ['recharge']:
            if all_case[i][1]=='login':
                result = qcd_requests(url, param, method, token=None)
                result_01=result
                token=result['data']['token_info']['token']
                token_1=token
                # print(token)
                print(result_01)
            else:
                result = qcd_requests(url, param, method, token='Bearer ' + token_1)  # 发起充值请求
                result_01=result
                print(result)
        from openpyxl import load_workbook
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(row=i+row,column=column_1).value=str(result_01)
        expect=eval(all_case[i][6])
        actual={'code':result_01['code'],'msg':result_01['msg']}
        if expect==actual:
            sheet.cell(row=i+row, column=column_2).value = 'pass'
            print('测试通过')
        else:
            sheet.cell(row=i+row, column=column_2).value = 'fail'
            print('测试不通过')
        wb.save(file_name)
run_write('qcd.test_case.xlsx','recharge',2,8,9)