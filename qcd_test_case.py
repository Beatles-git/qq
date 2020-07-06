import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="gb18030")
from openpyxl import load_workbook
def test_case(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    all_case=[]
    for i in range(1,sheet.max_row):
        case_column=[]
        for j in range(sheet.max_column):
            case=sheet.cell(row=i+1,column=j+1).value
            case_column.append(case)
        all_case.append(case_column)
        # print(case_column)
    # print(all_case)
    return all_case


# test_case('qcd.test_case.xlsx','login')
# print(all_case)